'''
Descripttion: THIS FILE IS PART OF Digital China PROJECT
version: 1.4
Author: Zhang BoQian (Daniel)
Date: 2023-03-23 9:16:50
LastEditors: error: error: git config user.name & please set dead value or install git && error: git config user.email & please set dead value or install git & please set dead value or install git
LastEditTime: 2023-04-04 16:02:36
# v1.1 版本说明
# 更新了IP地址获取错误问题
# 调整了生成的表格列宽
# 根据提取设备名称生成sheet页
# V1.2
# 增加收集信息失败日志记录
# 修复失败设备造成无法继续执行程序的问题
# V1.3 
# 更新表达式满足更多的设备信息位置
# v1.4
# 更新数据生成文件命名规则
# 更新文件保存位置
# 更新每个目录一个数据结果文件
'''
import os
import re
import openpyxl

# Define regular expressions to extract data from the text
device_id_pattern = re.compile(r"Device ID:\s*(\S+)")
local_intf_pattern = re.compile(r"Interface:\s*(\S+),\s*Port ID \(outgoing port\):\s*(\S+)")
platform_pattern = re.compile(r"Platform:\s*([\w\-]+)(?: \((.*)\))?")
ip_pattern = re.compile(r"(?:IPv4 Address|IP address):\s*(\S+)")
mgmt_ip_pattern = re.compile(r"(?:Mgmt|Management) address\(es\):\s*[\n\s]*(?:IPv4 Address|IP address):\s*(\S+)")
alt_ip_pattern = re.compile(r"Interface address\(es\):\s*[\n\s]*(?:IPv4 Address|IP address):\s*(\S+)")
# Traverse all *.txt files in the directory and its subdirectories

dir_path = r"C:\zbq\config"

for root, dirs, files in os.walk(dir_path):
    # Split the directory path into directory name and parent directory name
    parent_dirname, dirname = os.path.split(root)

    # Define the Excel file name for this directory
    excel_file_name = f"{dirname}_device_info.xlsx"
    #excel_file_path = os.path.join(root, excel_file_name)
    excel_file_path = os.path.join("C:\zbq\config", excel_file_name)
    # Create a new Excel workbook for this directory
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    skip_sheet = workbook.create_sheet(title="读取设备状态")
    skip_sheet.append(["读取设备", "读取状态"])
    skip_sheet.column_dimensions['A'].width = 60
    skip_sheet.column_dimensions['B'].width = 60

    # Process each text file in this directory
    for filename in files:
        if filename.endswith(".txt"):
            # Extract IP address from file name and use it as worksheet name
            ip_match = re.search(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", filename)
            if ip_match:
                ip = ip_match.group(0)
            else:
                skip_sheet.append([filename, "未获取到IP地址"])
                continue

            # Read the text file
            file_path = os.path.join(root, filename)
            with open(file_path, "r") as f:
                text = f.read()

            # Extract data from the text
            device_ids = device_id_pattern.findall(text)
            local_intfs = local_intf_pattern.findall(text)
            platforms = platform_pattern.findall(text)
            ips = ip_pattern.findall(text)
            mgmt_ips = mgmt_ip_pattern.findall(text)
            if not mgmt_ips:
                mgmt_ips = alt_ip_pattern.findall(text)

            # If no data was extracted, skip this file
            if not device_ids or not local_intfs or not platforms:
                skip_sheet.append([filename, "未获取到设备信息"])
                continue

            # Create worksheet
            worksheet = workbook.create_sheet(title=ip)

            # Write header
            worksheet.append(["Device ID", "Local Intrfce", "Platform", "Port ID", "IP"])

            # Write data
            for i in range(len(device_ids)):
                device_id = device_ids[i]
                local_intf = re.sub(r"FastEthernet(\d+)", r"Fas\1", local_intfs[i][0])
                platform = platforms[i][0] + (" (" + platforms[i][1] + ")" if platforms[i][1] else "")
                port_id = local_intfs[i][1]
                ip = mgmt_ips[i] if i < len(mgmt_ips) else ""
                worksheet.append([device_id, local_intf, platform, port_id, ip])

            # Set column width
            for column_letter in ['A', 'B', 'C', 'D', 'E']:
                    worksheet.column_dimensions[column_letter].width = 30

    # Save Excel file for this directory
    workbook.save(excel_file_path)